import io
import logging
import uuid
from datetime import datetime, timedelta, timezone

from dataclasses import dataclass
from typing import Literal

from openpyxl.styles import Alignment, Font
from sqlalchemy import func

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from openpyxl import Workbook

from app.bot.utils.excel import (
    format_worksheet,
    apply_table_style,
    merge_user_block,
    ColSpec,
    HEADER_FILL,
    apply_user_block_border,
    REJECTED_FILL,
    APPROVED_FILL,
)
from app.db.session import connection
from app.models import TaskAssignment, TaskReport, Task
from app.models.task_assignment import TaskAssignmentStatus
from app.models.user import User, UserApprovalStatus

logger = logging.getLogger(__name__)


MSC_TZ = timezone(timedelta(hours=3))

PeriodKey = Literal["day", "week", "all"]


@connection()
async def export_users_to_excel(*, session):
    stmt = (
        select(User)
        .options(
            selectinload(User.city),
            selectinload(User.referrer),
        )
        .order_by(User.id)
    )

    result = await session.execute(stmt)
    users = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    ws.append(
        [
            "Telegram ID (tg_id)",
            "Username (Telegram)",
            "ФИО пользователя",
            "Телефон",
            "Пол",
            "Город",
            "Реферер (ФИО + tg_id)",
        ]
    )

    for u in users:
        referrer = (
            f"{u.referrer.full_name or '—'} ({u.referrer.tg_id})" if u.referrer else "—"
        )

        ws.append(
            [
                u.tg_id,
                u.username or "—",
                u.full_name or "—",
                u.phone or "—",
                u.gender or "—",
                u.city.name if u.city else "—",
                referrer,
            ]
        )

    format_worksheet(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info("Экспорт пользователей в Excel: %s строк", len(users))
    return buffer


def _dt_to_ekb_str(dt: datetime | None) -> str:
    if not dt:
        return "—"
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(MSC_TZ).strftime("%Y-%m-%d %H:%M")


def gender_ru(value: str | None) -> str:
    return {
        "M": "Мужской",
        "F": "Женский",
    }.get(value, "—")


def assignment_status_ru(status) -> str:
    return {
        TaskAssignmentStatus.APPROVED: "Принято",
        TaskAssignmentStatus.REJECTED: "Отклонено",
        TaskAssignmentStatus.SUBMITTED: "На проверке",
        TaskAssignmentStatus.ASSIGNED: "Выдано",
    }.get(status, str(status))


@connection()
async def export_users_tasks_to_excel(
    *,
    session,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
) -> io.BytesIO:
    """
    Экспорт пользователей и их заданий.

    Особенности:
    - Верхняя строка с группировкой колонок
    - Переименованные колонки
    - Администратор выводится как ФИО @username (tg_id)
    - Аккаунт отчёта перенесён перед ссылкой
    - Включён autofilter
    """

    stmt = (
        select(TaskAssignment)
        .options(
            selectinload(TaskAssignment.user).selectinload(User.referrer),
            selectinload(TaskAssignment.user).selectinload(User.city),
            selectinload(TaskAssignment.task),
        )
        .order_by(
            TaskAssignment.user_id,
            TaskAssignment.processed_at.desc().nullslast(),
            TaskAssignment.submitted_at.desc().nullslast(),
        )
    )
    assignments = (await session.execute(stmt)).scalars().all()

    assignment_ids = [a.id for a in assignments]
    reports_map: dict[uuid.UUID, TaskReport] = {}

    if assignment_ids:
        rep_stmt = select(TaskReport).where(
            TaskReport.assignment_id.in_(assignment_ids)
        )
        reports = (await session.execute(rep_stmt)).scalars().all()
        reports_map = {r.assignment_id: r for r in reports}

    admin_ids = {
        a.processed_by_admin_id for a in assignments if a.processed_by_admin_id
    }

    admins: dict[int, User] = {}
    if admin_ids:
        stmt = select(User).where(User.tg_id.in_(admin_ids))
        admins = {u.tg_id: u for u in (await session.execute(stmt)).scalars().all()}

    by_user: dict[uuid.UUID, list[TaskAssignment]] = {}

    for a in assignments:
        if date_from or date_to:
            if a.status not in (
                TaskAssignmentStatus.APPROVED,
                TaskAssignmentStatus.REJECTED,
            ):
                continue
            if not a.processed_at:
                continue
            if date_from and a.processed_at < date_from:
                continue
            if date_to and a.processed_at >= date_to:
                continue

        by_user.setdefault(a.user_id, []).append(a)

    user_ids = set(by_user.keys())

    users_stmt = (
        select(User)
        .options(selectinload(User.city), selectinload(User.referrer))
        .order_by(User.id)
    )

    if date_from or date_to:
        users_stmt = users_stmt.where(User.id.in_(user_ids))

    users = (await session.execute(users_stmt)).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по заданиям"

    col_specs = [
        ColSpec("tg_id", "Telegram ID", 18),
        ColSpec("username", "@username", 22),
        ColSpec("full_name", "ФИО", 28),
        ColSpec("phone", "Телефон", 18),
        ColSpec("gender", "Пол", 10),
        ColSpec("city", "Город", 20),
        ColSpec("referrer", "Реферер (ФИО @username tg_id)", 36),
        ColSpec("status", "Статус задания", 18),
        ColSpec("report_account", "Аккаунт", 24),
        ColSpec("task_req_gender", "Требуемый пол", 14),
        ColSpec("task_req_city", "Требуемый город", 22),
        ColSpec("task_link", "Ссылка на задание", 36),
        ColSpec("task_example", "Пример", 50),
        ColSpec("submitted_at", "Отправлено (МСК)", 22),
        ColSpec("processed_at", "Проверено (МСК)", 22),
        ColSpec(
            "processed_by",
            "Принято администратором (ФИО @username tg_id)",
            42,
        ),
    ]

    USER_COLS_END = 7
    TOTAL_COLS = len(col_specs)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=USER_COLS_END)
    ws.cell(1, 1).value = "Информация о пользователе"

    ws.merge_cells(
        start_row=1, start_column=USER_COLS_END + 1, end_row=1, end_column=TOTAL_COLS
    )
    ws.cell(1, 8).value = "Отчёты"

    for col in (1, 8):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = HEADER_FILL

    ws.append([c.title for c in col_specs])
    WRAP_COLS = {
        "referrer",
        "processed_by",
    }
    for idx, spec in enumerate(col_specs, start=1):
        if spec.key in WRAP_COLS:
            ws.cell(row=2, column=idx).alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )
    current_row = 3

    for u in users:
        assignments = by_user.get(u.id, [])
        if not assignments:
            continue

        block_start = current_row

        referrer = (
            f"{u.referrer.full_name or '—'} "
            f"@{u.referrer.username or '—'} "
            f"({u.referrer.tg_id})"
            if u.referrer
            else "—"
        )

        for a in assignments:
            task = a.task

            req_gender = (
                "Мужской"
                if task.required_gender == "M"
                else "Женский"
                if task.required_gender == "F"
                else "—"
            )

            req_city = task.city.name if task.city else "Любой"

            report = reports_map.get(a.id)

            admin = admins.get(a.processed_by_admin_id)
            admin_str = (
                f"{admin.full_name or '—'} @{admin.username or '—'} ({admin.tg_id})"
                if admin
                else "—"
            )

            ws.append(
                [
                    u.tg_id,
                    f"@{u.username}" if u.username else "—",
                    u.full_name or "—",
                    u.phone or "—",
                    gender_ru(u.gender),
                    u.city.name if u.city else "—",
                    referrer,
                    assignment_status_ru(a.status),
                    report.account_name if report else "—",
                    req_gender,
                    req_city,
                    task.link if task else "—",
                    task.example_text if task else "—",
                    _dt_to_ekb_str(a.submitted_at),
                    _dt_to_ekb_str(a.processed_at),
                    admin_str,
                ]
            )

            status_cell = ws.cell(row=current_row, column=8)
            if a.status == TaskAssignmentStatus.APPROVED:
                status_cell.fill = APPROVED_FILL
            elif a.status == TaskAssignmentStatus.REJECTED:
                status_cell.fill = REJECTED_FILL

            current_row += 1

        merge_user_block(
            ws,
            start_row=block_start,
            end_row=current_row - 1,
            user_cols=[1, 2, 3, 4, 5, 6, 7],
        )
        apply_user_block_border(
            ws,
            start_row=block_start,
            end_row=current_row - 1,
            max_col=len(col_specs),
        )

    apply_table_style(ws, col_specs=col_specs)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


@dataclass(frozen=True)
class UserTaskItem:
    assignment_id: uuid.UUID
    status: str
    submitted_at: datetime | None
    processed_at: datetime | None
    processed_by_admin_id: int | None

    task_text: str | None
    task_example: str | None
    task_link: str | None


@connection()
async def get_user_by_tg_id(*, session, tg_id: int) -> User | None:
    stmt = (
        select(User)
        .options(selectinload(User.city), selectinload(User.referrer))
        .where(User.tg_id == tg_id)
        .limit(1)
    )
    return (await session.execute(stmt)).scalars().first()


def _period_to_range(period: PeriodKey) -> tuple[datetime | None, datetime | None]:
    """
    Возвращает (date_from, date_to) в часовом поясе MSC_TZ.
    - day: с начала текущего дня
    - week: с начала текущей недели (понедельник 00:00)
    - all: без ограничений
    """
    now = datetime.now(MSC_TZ)

    if period == "day":
        date_from = now.replace(hour=0, minute=0, second=0, microsecond=0)
        return date_from, None

    if period == "week":
        monday = now - timedelta(days=now.weekday())
        date_from = monday.replace(hour=0, minute=0, second=0, microsecond=0)
        return date_from, None

    return None, None


@connection()
async def get_user_tasks_page(
    *,
    session,
    tg_id: int,
    period: PeriodKey,
    page: int,
    page_size: int = 5,
) -> tuple[User | None, int, list[UserTaskItem]]:
    """
    Возвращает:
      - user (или None)
      - total_count (общее число заданий под фильтр)
      - items (страница заданий)
    Период применяется по processed_at (как в твоём экспорте) и только для APPROVED/REJECTED.
    Для period="all" — показываем все задания (в т.ч. ASSIGNED/SUBMITTED).
    """

    user = await get_user_by_tg_id(tg_id=tg_id)
    if not user:
        return None, 0, []

    date_from, date_to = _period_to_range(period)

    base_filters = [TaskAssignment.user_id == user.id]

    if period != "all":
        # как в экспорте: “за период” = обработанные админом
        base_filters.append(
            TaskAssignment.status.in_(
                [TaskAssignmentStatus.APPROVED, TaskAssignmentStatus.REJECTED]
            )
        )
        base_filters.append(TaskAssignment.processed_at.is_not(None))
        if date_from:
            base_filters.append(TaskAssignment.processed_at >= date_from)
        if date_to:
            base_filters.append(TaskAssignment.processed_at < date_to)

    # total
    total_stmt = select(func.count(TaskAssignment.id)).where(*base_filters)
    total_count = (await session.execute(total_stmt)).scalar_one()

    # page query
    offset = max(page, 0) * page_size

    stmt = (
        select(TaskAssignment)
        .options(
            selectinload(TaskAssignment.task),
        )
        .where(*base_filters)
        .order_by(
            TaskAssignment.processed_at.desc().nullslast(),
            TaskAssignment.submitted_at.desc().nullslast(),
            TaskAssignment.created_at.desc().nullslast(),
        )
        .offset(offset)
        .limit(page_size)
    )

    assignments = (await session.execute(stmt)).scalars().all()

    items: list[UserTaskItem] = []
    for a in assignments:
        t = a.task
        items.append(
            UserTaskItem(
                assignment_id=a.id,
                status=a.status,
                submitted_at=a.submitted_at,
                processed_at=a.processed_at,
                processed_by_admin_id=a.processed_by_admin_id,
                task_text=t.text if t else None,
                task_example=getattr(t, "example_text", None) if t else None,
                task_link=t.link if t else None,
            )
        )

    return user, int(total_count), items


@connection()
async def export_single_user_tasks_to_excel(
    *,
    session,
    tg_id: int,
    period: PeriodKey,
) -> io.BytesIO:
    """
    Excel экспорт заданий ОДНОГО пользователя по выбранному периоду.
    """
    user = await get_user_by_tg_id(tg_id=tg_id)
    if not user:
        # пустой файл
        wb = Workbook()
        ws = wb.active
        ws.title = "User_Tasks"
        ws.append(["Пользователь не найден"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    date_from, date_to = _period_to_range(period)

    base_filters = [TaskAssignment.user_id == user.id]

    if period != "all":
        base_filters.append(
            TaskAssignment.status.in_(
                [TaskAssignmentStatus.APPROVED, TaskAssignmentStatus.REJECTED]
            )
        )
        base_filters.append(TaskAssignment.processed_at.is_not(None))
        if date_from:
            base_filters.append(TaskAssignment.processed_at >= date_from)
        if date_to:
            base_filters.append(TaskAssignment.processed_at < date_to)

    stmt = (
        select(TaskAssignment)
        .options(selectinload(TaskAssignment.task))
        .where(*base_filters)
        .order_by(
            TaskAssignment.processed_at.desc().nullslast(),
            TaskAssignment.submitted_at.desc().nullslast(),
            TaskAssignment.created_at.desc().nullslast(),
        )
    )

    assignments = (await session.execute(stmt)).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "User_Tasks"

    col_specs = [
        ColSpec("tg_id", "tg_id", 18),
        ColSpec("full_name", "ФИО", 28),
        ColSpec("phone", "Телефон", 18),
        ColSpec("gender", "Пол", 10),
        ColSpec("city", "Город", 20),
        ColSpec("status", "Статус", 18),
        ColSpec("submitted_at", "Отправлено (ЕКБ)", 22),
        ColSpec("processed_at", "Проверено (ЕКБ)", 22),
        ColSpec("processed_by", "admin tg_id", 16),
        ColSpec("task_link", "Ссылка", 36),
        ColSpec("task_text", "Текст", 50),
        ColSpec("task_example", "Пример", 50),
    ]

    ws.append([c.title for c in col_specs])

    for a in assignments:
        t = a.task
        ws.append(
            [
                user.tg_id,
                user.full_name or "—",
                user.phone or "—",
                user.gender or "—",
                user.city.name if user.city else "—",
                a.status,
                _dt_to_ekb_str(a.submitted_at),
                _dt_to_ekb_str(a.processed_at),
                str(a.processed_by_admin_id) if a.processed_by_admin_id else "—",
                t.link if t else "—",
                t.text if t else "—",
                getattr(t, "example_text", None) or "—",
            ]
        )

    apply_table_style(ws, col_specs=col_specs)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@connection()
async def set_user_blocked(
    *,
    session,
    tg_id: int,
    blocked: bool,
) -> None:
    result = await session.execute(select(User).where(User.tg_id == tg_id))
    user = result.scalar_one_or_none()
    if not user:
        return

    user.is_blocked = blocked
    user.blocked_at = datetime.now(timezone.utc) if blocked else None

    await session.commit()


@connection()
async def get_daily_completed_stats(*, session):
    """
    Возвращает количество APPROVED заданий по дням за последние 7 дней.
    """
    now = datetime.now(MSC_TZ)
    date_from = (now - timedelta(days=6)).replace(
        hour=0, minute=0, second=0, microsecond=0
    )

    stmt = (
        select(
            func.date(TaskAssignment.processed_at).label("day"),
            func.count(TaskAssignment.id),
        )
        .where(
            TaskAssignment.status == TaskAssignmentStatus.APPROVED,
            TaskAssignment.processed_at >= date_from,
        )
        .group_by("day")
        .order_by("day")
    )

    rows = (await session.execute(stmt)).all()

    stats_map = {row[0]: row[1] for row in rows}

    result = []
    for i in range(7):
        day = (date_from + timedelta(days=i)).date()
        count = stats_map.get(day, 0)
        result.append((day.strftime("%d.%m"), count))

    return result


@connection()
async def get_top_5_users(*, session):
    """
    Возвращает топ-5 пользователей по количеству APPROVED заданий.
    """
    stmt = (
        select(
            User.id,
            User.full_name,
            User.tg_id,
            User.username,
            func.count(TaskAssignment.id).label("approved_count"),
        )
        .join(TaskAssignment, TaskAssignment.user_id == User.id)
        .where(
            TaskAssignment.status == TaskAssignmentStatus.APPROVED,
            User.is_blocked.is_(False),
        )
        .group_by(
            User.id,
            User.full_name,
            User.tg_id,
            User.username,
        )
        .order_by(func.count(TaskAssignment.id).desc())
        .limit(5)
    )

    rows = (await session.execute(stmt)).all()

    return [
        {
            "id": row[0],
            "name": row[1] or "—",
            "tg_id": row[2],
            "username": row[3],
            "count": row[4],
        }
        for row in rows
    ]


def _dt_to_msk_str(dt):
    if not dt:
        return "—"
    return dt.astimezone(MSC_TZ).strftime("%Y-%m-%d %H:%M")


@connection()
async def export_available_tasks_to_excel(*, session) -> io.BytesIO:
    """
    Excel-экспорт доступных заданий
    """

    taken_subquery = select(TaskAssignment.task_id).where(
        TaskAssignment.is_archived.is_(False)
    )

    stmt = (
        select(Task)
        .options(selectinload(Task.city))
        .where(~Task.id.in_(taken_subquery))
        .order_by(Task.created_at.desc())
    )

    tasks = (await session.execute(stmt)).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Available_Tasks"

    col_specs = [
        ColSpec("created_at", "Создано (МСК)", 22),
        ColSpec("source", "Источник", 18),
        ColSpec("city", "Город", 20),
        ColSpec("required_gender", "От какого лица", 14),
        ColSpec("link", "Ссылка", 40),
        ColSpec("text", "Текст задания", 55),
        ColSpec("example_text", "Пример отзыва", 55),
    ]
    ws.append(["Доступные задания"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_specs))
    ws["A1"].font = Font(bold=True)

    ws.append([c.title for c in col_specs])

    for t in tasks:
        ws.append(
            [
                _dt_to_msk_str(t.created_at),
                t.source or "—",
                t.city.name if t.city else "-",
                gender_ru(t.required_gender),
                t.link,
                t.text,
                t.example_text or "—",
            ]
        )

    apply_table_style(ws, col_specs=col_specs)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


@connection()
async def get_users_statistics(*, session: AsyncSession) -> dict:
    now = datetime.now(MSC_TZ)

    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=7)
    month_start = today_start - timedelta(days=30)

    total_users = await session.scalar(
        select(func.count(User.id)).where(
            User.approval_status == UserApprovalStatus.APPROVED
        )
    )

    new_today = await session.scalar(
        select(func.count(User.id)).where(
            User.approval_status == UserApprovalStatus.APPROVED,
            User.approval_at >= today_start
        )
    )

    new_week = await session.scalar(
        select(func.count(User.id)).where(
            User.approval_status == UserApprovalStatus.APPROVED,
            User.approval_at >= week_start
        )
    )

    new_month = await session.scalar(
        select(func.count(User.id)).where(
            User.approval_status == UserApprovalStatus.APPROVED,
            User.approval_at >= month_start
        )
    )

    return {
        "total_users": total_users or 0,
        "new_today": new_today or 0,
        "new_week": new_week or 0,
        "new_month": new_month or 0,
    }


@connection()
async def get_user_weekly_approved_count(*, session, user_id):
    week_start = datetime.now(MSC_TZ) - timedelta(days=7)

    stmt = select(func.count(TaskAssignment.id)).where(
        TaskAssignment.user_id == user_id,
        TaskAssignment.status == TaskAssignmentStatus.APPROVED,
        TaskAssignment.processed_at >= week_start,
    )

    return await session.scalar(stmt) or 0
