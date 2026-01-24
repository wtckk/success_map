from dataclasses import dataclass
from typing import Iterable

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="ECECEC", end_color="ECECEC", fill_type="solid")
FIRST_COL_FILL = PatternFill(
    start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
)
APPROVED_FILL = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid",
)

REJECTED_FILL = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid",
)


def format_worksheet(ws: Worksheet) -> None:
    """
    Приводит Excel-лист в читаемый вид:
    - жирная шапка
    - автоширина колонок
    - первый столбец шире и выделен
    """

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    first_col = get_column_letter(1)
    ws.column_dimensions[first_col].width = 22

    for cell in ws[first_col]:
        cell.fill = FIRST_COL_FILL

    ws.freeze_panes = "A2"


THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

THICK = Side(style="medium")
USER_BORDER = Border(bottom=THICK)


@dataclass(frozen=True)
class ColSpec:
    key: str
    title: str
    width: int


def apply_user_block_border(
    ws: Worksheet,
    *,
    start_row: int,
    end_row: int,
    max_col: int,
) -> None:
    """
    Добавляет жирную нижнюю границу всему блоку пользователя.
    """
    for col in range(1, max_col + 1):
        cell = ws.cell(row=end_row, column=col)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top,
            bottom=THICK,
        )


def apply_table_style(ws: Worksheet, *, col_specs: list[ColSpec]) -> None:
    HEADER_ROW = 2

    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        cell = ws.cell(row=HEADER_ROW, column=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for idx, spec in enumerate(col_specs, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = spec.width

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 18)
    for r in range(1, max_row + 1):
        ws.cell(row=r, column=1).fill = FIRST_COL_FILL

    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = 28
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(max_col)}{max_row}"


def merge_user_block(
    ws: Worksheet, *, start_row: int, end_row: int, user_cols: Iterable[int]
) -> None:
    """
    Объединяем ячейки пользователя по высоте (вертикально) для набора колонок.
    """
    if end_row <= start_row:
        return

    for col_idx in user_cols:
        ws.merge_cells(
            start_row=start_row,
            start_column=col_idx,
            end_row=end_row,
            end_column=col_idx,
        )
        cell = ws.cell(row=start_row, column=col_idx)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
